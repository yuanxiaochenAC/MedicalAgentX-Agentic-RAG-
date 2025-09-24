import os
import json
import pickle
import csv
import yaml
import xml.etree.ElementTree as ET
from typing import Dict, Any, List
from pathlib import Path
import mimetypes
import hashlib
from datetime import datetime

# For handling various file types
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from PIL import Image
    PILLOW_AVAILABLE = True
except ImportError:
    PILLOW_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from ..core.module import BaseModule
from ..core.logging import logger


class StorageBase(BaseModule):
    """
    Base class for comprehensive storage operations supporting various file types.
    Provides save, read, and append functionality for documents, data files, images, videos, and sound files.
    Designed with scalability in mind for future database integration.
    """
    
    def __init__(self, base_path: str = ".", **kwargs):
        """
        Initialize the StorageBase with configuration options.
        
        Args:
            base_path (str): Base directory for storage operations (default: current directory)
            **kwargs: Additional keyword arguments for parent class initialization
        """
        super().__init__(**kwargs)
        self.base_path = Path(base_path)
        self.base_path.mkdir(parents=True, exist_ok=True)
        
        # File types that support append operations
        self.appendable_formats = {
            '.txt': self._append_text,
            '.json': self._append_json,
            '.csv': self._append_csv,
            '.yaml': self._append_yaml,
            '.yml': self._append_yaml,
            '.pickle': self._append_pickle,
            '.xlsx': self._append_excel
        }
    
    def _resolve_path(self, file_path: str) -> Path:
        """Resolve file path, prepending base_path if it's a relative path"""
        path = Path(file_path)
        if not path.is_absolute():
            # If it's a relative path, prepend the base path
            path = self.base_path / path
        return path
    
    def get_file_type(self, file_path: str) -> str:
        """Get the file extension from a file path"""
        return Path(file_path).suffix.lower()
    
    def get_mime_type(self, file_path: str) -> str:
        """Get the MIME type of a file"""
        return mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get comprehensive information about a file"""
        try:
            path = self._resolve_path(file_path)
            if not path.exists():
                return {"success": False, "error": f"File {file_path} does not exist"}
            
            stat = path.stat()
            return {
                "success": True,
                "file_path": str(path),
                "file_name": path.name,
                "file_extension": path.suffix.lower(),
                "mime_type": self.get_mime_type(str(path)),
                "size_bytes": stat.st_size,
                "size_mb": round(stat.st_size / (1024 * 1024), 2),
                "created_time": datetime.fromtimestamp(stat.st_ctime).isoformat(),
                "modified_time": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                "is_file": path.is_file(),
                "is_directory": path.is_dir(),
                "is_readable": os.access(path, os.R_OK),
                "is_writable": os.access(path, os.W_OK)
            }
        except Exception as e:
            logger.error(f"Error getting file info for {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def calculate_file_hash(self, file_path: str, algorithm: str = 'md5') -> Dict[str, Any]:
        """Calculate hash of a file"""
        try:
            resolved_path = self._resolve_path(file_path)
            hash_func = hashlib.new(algorithm)
            with open(resolved_path, 'rb') as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_func.update(chunk)
            
            return {
                "success": True,
                "file_path": str(resolved_path),
                "algorithm": algorithm,
                "hash": hash_func.hexdigest()
            }
        except Exception as e:
            logger.error(f"Error calculating hash for {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def save(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Save content to a file with automatic format detection"""
        try:
            # Resolve the file path (prepend base_path if relative)
            resolved_path = self._resolve_path(file_path)
            file_extension = resolved_path.suffix.lower()
            
            # Ensure directory exists
            resolved_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Handle different file types
            if file_extension == '.json':
                return self._save_json(str(resolved_path), content, **kwargs)
            elif file_extension in ['.yaml', '.yml']:
                return self._save_yaml(str(resolved_path), content, **kwargs)
            elif file_extension == '.csv':
                return self._save_csv(str(resolved_path), content, **kwargs)
            elif file_extension == '.xlsx':
                return self._save_excel(str(resolved_path), content, **kwargs)
            elif file_extension == '.xml':
                return self._save_xml(str(resolved_path), content, **kwargs)
            elif file_extension == '.pickle':
                return self._save_pickle(str(resolved_path), content, **kwargs)
            elif file_extension == '.pdf':
                return self._save_pdf(str(resolved_path), content, **kwargs)
            elif file_extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']:
                return self._save_image(str(resolved_path), content, **kwargs)
            else:
                # Default to text/binary handling
                return self._save_text(str(resolved_path), content, **kwargs)
            
        except Exception as e:
            logger.error(f"Error saving file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": str(file_path)}
    
    def read(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read content from a file with automatic format detection"""
        try:
            # Resolve the file path (prepend base_path if relative)
            resolved_path = self._resolve_path(file_path)
            if not resolved_path.exists():
                return {"success": False, "error": f"File {file_path} does not exist"}
            
            file_extension = resolved_path.suffix.lower()
            
            # Handle different file types
            if file_extension == '.json':
                return self._read_json(str(resolved_path), **kwargs)
            elif file_extension in ['.yaml', '.yml']:
                return self._read_yaml(str(resolved_path), **kwargs)
            elif file_extension == '.csv':
                return self._read_csv(str(resolved_path), **kwargs)
            elif file_extension == '.xlsx':
                return self._read_excel(str(resolved_path), **kwargs)
            elif file_extension == '.xml':
                return self._read_xml(str(resolved_path), **kwargs)
            elif file_extension == '.pickle':
                return self._read_pickle(str(resolved_path), **kwargs)
            elif file_extension == '.pdf':
                return self._read_pdf(str(resolved_path), **kwargs)
            elif file_extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']:
                return self._read_image(str(resolved_path), **kwargs)
            else:
                # Default to text/binary handling
                return self._read_text(str(resolved_path), **kwargs)
            
        except Exception as e:
            logger.error(f"Error reading file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": str(file_path)}
    
    def append(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Append content to a file (only for supported formats)"""
        try:
            # Resolve the file path (prepend base_path if relative)
            resolved_path = self._resolve_path(file_path)
            file_extension = resolved_path.suffix.lower()
            
            # Check if format supports append
            if file_extension not in self.appendable_formats:
                return {
                    "success": False, 
                    "error": f"Append not supported for {file_extension} files",
                    "supported_formats": list(self.appendable_formats.keys())
                }
            
            # Ensure directory exists
            resolved_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Use the appropriate append handler
            append_handler = self.appendable_formats[file_extension]
            return append_handler(str(resolved_path), content, **kwargs)
            
        except Exception as e:
            logger.error(f"Error appending to file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": str(file_path)}
    
    def get_supported_formats(self) -> Dict[str, Any]:
        """Get information about supported file formats"""
        return {
            "success": True,
            "appendable_formats": list(self.appendable_formats.keys()),
            "all_formats": [
                '.txt', '.json', '.csv', '.yaml', '.yml', '.xml', '.xlsx', '.pickle',
                '.pdf', '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff',
                '.mp4', '.avi', '.mp3', '.wav'
            ],
            "categories": {
                "Documents": ['.pdf', '.txt'],
                "Data Files": ['.json', '.csv', '.yaml', '.yml', '.xml', '.xlsx', '.pickle'],
                "Images": ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff'],
                "Media": ['.mp4', '.avi', '.mp3', '.wav']
            }
        }
    
    def list_files(self, path: str = None, max_depth: int = 3, include_hidden: bool = False) -> Dict[str, Any]:
        """List files and directories in a structured tree format"""
        try:
            if path is None:
                path = str(self.base_path)
            
            path_obj = Path(path)
            if not path_obj.exists():
                return {"success": False, "error": f"Path {path} does not exist"}
            
            if not path_obj.is_dir():
                return {"success": False, "error": f"Path {path} is not a directory"}
            
            def build_tree(current_path: Path, current_depth: int) -> Dict[str, Any]:
                if current_depth > max_depth:
                    return {"type": "truncated", "path": str(current_path)}
                
                result = {
                    "type": "directory",
                    "path": str(current_path),
                    "name": current_path.name,
                    "items": []
                }
                
                try:
                    items = list(current_path.iterdir())
                    
                    # Filter hidden files if needed
                    if not include_hidden:
                        items = [item for item in items if not item.name.startswith('.')]
                    
                    # Sort: directories first, then files
                    directories = [item for item in items if item.is_dir()]
                    files = [item for item in items if item.is_file()]
                    
                    directories.sort(key=lambda x: x.name.lower())
                    files.sort(key=lambda x: x.name.lower())
                    
                    # Add directories
                    for item in directories:
                        if current_depth < max_depth:
                            result["items"].append(build_tree(item, current_depth + 1))
                        else:
                            result["items"].append({
                                "type": "directory",
                                "path": str(item),
                                "name": item.name,
                                "items": []
                            })
                    
                    # Add files
                    for item in files:
                        try:
                            stat = item.stat()
                            file_info = {
                                "type": "file",
                                "path": str(item),
                                "name": item.name,
                                "extension": item.suffix.lower(),
                                "size_bytes": stat.st_size,
                                "size_mb": round(stat.st_size / (1024 * 1024), 2),
                                "modified_time": datetime.fromtimestamp(stat.st_mtime).isoformat()
                            }
                            result["items"].append(file_info)
                        except Exception:
                            # Skip files we can't access
                            continue
                            
                except PermissionError:
                    result["error"] = "Permission denied"
                except Exception as e:
                    result["error"] = str(e)
                
                return result
            
            tree = build_tree(path_obj, 0)
            
            # Add summary statistics
            def count_items(node: Dict[str, Any]) -> tuple:
                files = 0
                dirs = 0
                total_size = 0
                
                if node.get("type") == "file":
                    files = 1
                    total_size = node.get("size_bytes", 0)
                elif node.get("type") == "directory":
                    dirs = 1
                    for item in node.get("items", []):
                        f, d, s = count_items(item)
                        files += f
                        dirs += d
                        total_size += s
                
                return files, dirs, total_size
            
            files_count, dirs_count, total_size = count_items(tree)
            
            return {
                "success": True,
                "items": tree.get("items", []),
                "tree": tree,
                "summary": {
                    "path": str(path_obj),
                    "total_files": files_count,
                    "total_directories": dirs_count,
                    "total_size_bytes": total_size,
                    "total_size_mb": round(total_size / (1024 * 1024), 2),
                    "max_depth": max_depth,
                    "include_hidden": include_hidden
                }
            }
            
        except Exception as e:
            logger.error(f"Error listing files in {path}: {str(e)}")
            return {"success": False, "error": str(e), "path": path}
    
    # Text file handlers
    def _save_text(self, file_path: str, content: Any, encoding: str = 'utf-8', **kwargs) -> Dict[str, Any]:
        """Save text content to a file"""
        try:
            with open(file_path, 'w', encoding=encoding) as f:
                f.write(str(content))
            
            return {
                "success": True,
                "message": f"File saved to {file_path}",
                "file_path": file_path,
                "content_length": len(str(content))
            }
        except Exception as e:
            logger.error(f"Error saving text file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_text(self, file_path: str, encoding: str = 'utf-8', **kwargs) -> Dict[str, Any]:
        """Read text content from a file"""
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path,
                "content_length": len(content)
            }
        except Exception as e:
            logger.error(f"Error reading text file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # JSON file handlers
    def _save_json(self, file_path: str, content: Any, indent: int = 2, **kwargs) -> Dict[str, Any]:
        """Save JSON content to a file"""
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(content, f, indent=indent, ensure_ascii=False)
            
            return {
                "success": True,
                "message": f"JSON file saved to {file_path}",
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error saving JSON file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_json(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read JSON content from a file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = json.load(f)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error reading JSON file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_json(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Append content to JSON file (for arrays)"""
        try:
            existing_content = []
            if Path(file_path).exists():
                with open(file_path, 'r', encoding='utf-8') as f:
                    existing_content = json.load(f)
            
            if isinstance(existing_content, list):
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            elif isinstance(existing_content, dict):
                # If existing content is a dict, merge with new content
                if isinstance(content, dict):
                    existing_content.update(content)
                else:
                    return {"success": False, "error": "Cannot append non-dict to JSON dict"}
            else:
                # Convert to list and append
                existing_content = [existing_content]
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(existing_content, f, indent=2, ensure_ascii=False)
            
            return {
                "success": True,
                "message": f"Content appended to JSON file {file_path}",
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error appending to JSON file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # CSV file handlers
    def _save_csv(self, file_path: str, content: List[Dict[str, Any]], **kwargs) -> Dict[str, Any]:
        """Save CSV content to a file"""
        try:
            if not content:
                return {"success": False, "error": "No content to save"}
            
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                fieldnames = content[0].keys()
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(content)
            
            return {
                "success": True,
                "message": f"CSV file saved to {file_path}",
                "file_path": file_path,
                "rows": len(content)
            }
        except Exception as e:
            logger.error(f"Error saving CSV file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_csv(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read CSV content from a file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                content = list(reader)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path,
                "rows": len(content)
            }
        except Exception as e:
            logger.error(f"Error reading CSV file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_csv(self, file_path: str, content: List[Dict[str, Any]], **kwargs) -> Dict[str, Any]:
        """Append content to CSV file"""
        try:
            with open(file_path, 'a', newline='', encoding='utf-8') as f:
                if not content:
                    return {"success": False, "error": "No content to append"}
                
                fieldnames = content[0].keys()
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writerows(content)
            
            return {
                "success": True,
                "message": f"Content appended to CSV file {file_path}",
                "file_path": file_path,
                "appended_rows": len(content)
            }
        except Exception as e:
            logger.error(f"Error appending to CSV file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # YAML file handlers
    def _save_yaml(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Save YAML content to a file"""
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                yaml.dump(content, f, default_flow_style=False, allow_unicode=True)
            
            return {
                "success": True,
                "message": f"YAML file saved to {file_path}",
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error saving YAML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_yaml(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read YAML content from a file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = yaml.safe_load(f)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error reading YAML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_yaml(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Append content to YAML file (for lists)"""
        try:
            existing_content = []
            if Path(file_path).exists():
                with open(file_path, 'r', encoding='utf-8') as f:
                    existing_content = yaml.safe_load(f) or []
            
            if isinstance(existing_content, list):
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            elif isinstance(existing_content, dict):
                # If existing content is a dict, merge with new content
                if isinstance(content, dict):
                    existing_content.update(content)
                else:
                    return {"success": False, "error": "Cannot append non-dict to YAML dict"}
            else:
                # Convert to list and append
                existing_content = [existing_content]
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                yaml.dump(existing_content, f, default_flow_style=False, allow_unicode=True)
            
            return {
                "success": True,
                "message": f"Content appended to YAML file {file_path}",
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error appending to YAML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # XML file handlers
    def _save_xml(self, file_path: str, content: Any, root_tag: str = "root", **kwargs) -> Dict[str, Any]:
        """Save XML content to a file"""
        try:
            # If content is already a string (raw XML), write it directly
            if isinstance(content, str):
                # Check if it's already valid XML
                try:
                    ET.fromstring(content)
                    # It's valid XML, write it directly
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(content)
                    return {
                        "success": True,
                        "message": f"XML file saved to {file_path}",
                        "file_path": file_path
                    }
                except ET.ParseError:
                    # Not valid XML, treat as text content and wrap it
                    pass
            
            # If content is a dictionary, convert to XML
            if isinstance(content, dict):
                def dict_to_xml(data, root):
                    for key, value in data.items():
                        child = ET.SubElement(root, key)
                        if isinstance(value, dict):
                            dict_to_xml(value, child)
                        else:
                            child.text = str(value)
                
                root = ET.Element(root_tag)
                dict_to_xml(content, root)
                tree = ET.ElementTree(root)
                tree.write(file_path, encoding='utf-8', xml_declaration=True)
            else:
                # For other types, wrap in root element
                root = ET.Element(root_tag)
                root.text = str(content)
                tree = ET.ElementTree(root)
                tree.write(file_path, encoding='utf-8', xml_declaration=True)
            
            return {
                "success": True,
                "message": f"XML file saved to {file_path}",
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error saving XML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_xml(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read XML content from a file"""
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            def xml_to_dict(element):
                result = {}
                for child in element:
                    if len(child) == 0:
                        result[child.tag] = child.text
                    else:
                        result[child.tag] = xml_to_dict(child)
                return result
            
            content = xml_to_dict(root)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error reading XML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Excel file handlers
    def _save_excel(self, file_path: str, content: List[List[Any]], sheet_name: str = "Sheet1", **kwargs) -> Dict[str, Any]:
        """Save Excel content to a file"""
        if not EXCEL_AVAILABLE:
            return {"success": False, "error": "openpyxl library not available"}
        
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            
            for row in content:
                worksheet.append(row)
            
            workbook.save(file_path)
            
            return {
                "success": True,
                "message": f"Excel file saved to {file_path}",
                "file_path": file_path,
                "rows": len(content)
            }
        except Exception as e:
            logger.error(f"Error saving Excel file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_excel(self, file_path: str, sheet_name: str = None, **kwargs) -> Dict[str, Any]:
        """Read Excel content from a file"""
        if not EXCEL_AVAILABLE:
            return {"success": False, "error": "openpyxl library not available"}
        
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet_names = workbook.sheetnames
            
            if sheet_name is None:
                sheet_name = sheet_names[0]
            
            if sheet_name not in sheet_names:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            worksheet = workbook[sheet_name]
            content = []
            
            for row in worksheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    content.append(list(row))
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "rows": len(content)
            }
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_excel(self, file_path: str, content: List[List[Any]], sheet_name: str = None, **kwargs) -> Dict[str, Any]:
        """Append content to Excel file"""
        if not EXCEL_AVAILABLE:
            return {"success": False, "error": "openpyxl library not available"}
        
        try:
            if not Path(file_path).exists():
                return self._save_excel(file_path, content, sheet_name or "Sheet1", **kwargs)
            
            workbook = load_workbook(file_path)
            sheet_names = workbook.sheetnames
            
            if sheet_name is None:
                sheet_name = sheet_names[0]
            
            if sheet_name not in sheet_names:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            worksheet = workbook[sheet_name]
            
            for row in content:
                worksheet.append(row)
            
            workbook.save(file_path)
            
            return {
                "success": True,
                "message": f"Content appended to Excel file {file_path}",
                "file_path": file_path,
                "appended_rows": len(content)
            }
        except Exception as e:
            logger.error(f"Error appending to Excel file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Pickle file handlers
    def _save_pickle(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Save pickle content to a file"""
        try:
            with open(file_path, 'wb') as f:
                pickle.dump(content, f)
            
            return {
                "success": True,
                "message": f"Pickle file saved to {file_path}",
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error saving pickle file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_pickle(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read pickle content from a file"""
        try:
            with open(file_path, 'rb') as f:
                content = pickle.load(f)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error reading pickle file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_pickle(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Append content to pickle file (for lists)"""
        try:
            existing_content = []
            if Path(file_path).exists():
                with open(file_path, 'rb') as f:
                    existing_content = pickle.load(f)
            
            if isinstance(existing_content, list):
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            elif isinstance(existing_content, dict):
                # If existing content is a dict, try to merge intelligently
                if isinstance(content, dict):
                    existing_content.update(content)
                elif isinstance(content, list):
                    # If appending a list to a dict, add it as a new key
                    existing_content["appended_list"] = content
                else:
                    # If appending a single value to a dict, add it as a new key
                    existing_content["appended_value"] = content
            else:
                # Convert to list and append
                existing_content = [existing_content]
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            
            with open(file_path, 'wb') as f:
                pickle.dump(existing_content, f)
            
            return {
                "success": True,
                "message": f"Content appended to pickle file {file_path}",
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error appending to pickle file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # PDF file handlers
    def _save_pdf(self, file_path: str, content: str, **kwargs) -> Dict[str, Any]:
        """Save content to a PDF file"""
        try:
            # Use reportlab to create a proper PDF with text content
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Split content into paragraphs
            paragraphs = content.split('\n')
            
            for para_text in paragraphs:
                if para_text.strip():  # Non-empty paragraph
                    para = Paragraph(para_text, styles['Normal'])
                    story.append(para)
                    story.append(Spacer(1, 12))
                else:
                    story.append(Spacer(1, 12))
            
            # Build the PDF
            doc.build(story)
            
            return {
                "success": True,
                "message": f"PDF file saved to {file_path}",
                "file_path": file_path
            }
                
        except ImportError:
            return {"success": False, "error": "reportlab library not available for PDF creation"}
        except Exception as e:
            logger.error(f"Error saving PDF file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_pdf(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read content from a PDF file"""
        if not PDF_AVAILABLE:
            return {"success": False, "error": "PyPDF2 library not available"}
        
        try:
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                text = ""
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text += page.extract_text() + "\n"
                
                return {
                    "success": True,
                    "content": text,
                    "file_path": file_path,
                    "pages": len(pdf_reader.pages)
                }
                
        except Exception as e:
            logger.error(f"Error reading PDF file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Image file handlers
    def _save_image(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Save image content to a file"""
        if not PILLOW_AVAILABLE:
            return {"success": False, "error": "Pillow library not available"}
        
        try:
            # Check if content is already a PIL Image object
            if hasattr(content, 'save') and callable(getattr(content, 'save', None)):
                # Content is a PIL Image object
                content.save(file_path)
                return {
                    "success": True,
                    "message": f"Image saved to {file_path}",
                    "file_path": file_path,
                    "format": content.format,
                    "size": content.size
                }
            elif isinstance(content, bytes):
                # Content is binary image data
                with open(file_path, 'wb') as f:
                    f.write(content)
                return {
                    "success": True,
                    "message": f"Image saved to {file_path}",
                    "file_path": file_path
                }
            elif isinstance(content, str) and Path(content).exists():
                # Content is a file path to an existing image
                import shutil
                shutil.copy2(content, file_path)
                return {
                    "success": True,
                    "message": f"Image copied from {content} to {file_path}",
                    "file_path": file_path
                }
            else:
                return {"success": False, "error": "Content must be a PIL Image object, binary data, or valid file path"}
                
        except Exception as e:
            logger.error(f"Error saving image file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_image(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read image and return PIL Image object"""
        if not PILLOW_AVAILABLE:
            return {"success": False, "error": "Pillow library not available"}
        
        try:
            with Image.open(file_path) as img:
                # Convert to RGB if necessary for consistency
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                
                metadata = {
                    "format": img.format,
                    "mode": img.mode,
                    "size": img.size,
                    "width": img.width,
                    "height": img.height
                }
                
                return {
                    "success": True,
                    "content": img,  # Return the PIL Image object
                    "metadata": metadata,
                    "file_path": file_path
                }
                
        except Exception as e:
            logger.error(f"Error reading image file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Text append handler
    def _append_text(self, file_path: str, content: str, encoding: str = 'utf-8', **kwargs) -> Dict[str, Any]:
        """Append text content to a file"""
        try:
            with open(file_path, 'a', encoding=encoding) as f:
                f.write(str(content))
            
            return {
                "success": True,
                "message": f"Content appended to file {file_path}",
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error appending to text file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Placeholder for future database integration
    def _get_database_connection(self, db_type: str, connection_string: str) -> Any:
        """Placeholder for future database integration"""
        # This will be implemented when adding database support
        raise NotImplementedError("Database integration not yet implemented") 